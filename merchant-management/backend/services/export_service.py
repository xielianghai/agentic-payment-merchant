from datetime import datetime, timezone
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from services.audit_service import log_operation, list_operation_logs
from services.common import dumps, loads, one, rows
from services.reconciliation_service import list_items, list_runs
from services.transaction_service import list_transactions


def _serialize_row(row: dict[str, Any]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for key, value in row.items():
        if isinstance(value, datetime):
            out[key] = value.isoformat()
        elif isinstance(value, Decimal):
            out[key] = float(value)
        elif isinstance(value, dict):
            out[key] = _serialize_row(value)
        elif isinstance(value, list):
            out[key] = [_serialize_row(v) if isinstance(v, dict) else v for v in value]
        else:
            out[key] = value
    return out


def _cell(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, dict):
        return dumps(value)
    return value


def _write_sheet(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append([_cell(v) for v in row])


async def _gather_dispute_bundle(session: AsyncSession, merchant_id: str) -> dict[str, Any]:
    transactions = await list_transactions(session, merchant_id=merchant_id, limit=500)
    logs = await list_operation_logs(session, merchant_id=merchant_id, limit=100)
    runs = await list_runs(session, merchant_id)
    reconciliation_items: list[dict[str, Any]] = []
    if runs:
        reconciliation_items = await list_items(session, runs[0]["id"])
    return {
        "transactions": transactions,
        "logs": logs,
        "runs": runs,
        "reconciliation_items": reconciliation_items,
    }


def build_dispute_excel_bytes(
    merchant_id: str,
    bundle: dict[str, Any],
    generated_at: datetime,
) -> bytes:
    transactions: list[dict[str, Any]] = bundle["transactions"]
    logs: list[dict[str, Any]] = bundle["logs"]
    runs: list[dict[str, Any]] = bundle["runs"]
    reconciliation_items: list[dict[str, Any]] = bundle["reconciliation_items"]
    latest_run = runs[0] if runs else None
    failed_txs = [t for t in transactions if t.get("status") == "FAILED"]
    mandate_fail = [
        t for t in failed_txs if (t.get("detail_json") or {}).get("error") == "mandate_verify_fail"
    ]

    wb = Workbook()
    summary = wb.active
    summary.title = "Summary"
    summary.append(["Field", "Value"])
    summary.append(["merchant_id", merchant_id])
    summary.append(["generated_at", generated_at.isoformat(sep=" ", timespec="seconds")])
    summary.append(["transaction_count", len(transactions)])
    summary.append(["failed_transaction_count", len(failed_txs)])
    summary.append(["mandate_verify_fail_count", len(mandate_fail)])
    summary.append(["reconciliation_run_count", len(runs)])
    summary.append(["latest_reconciliation_run_id", latest_run["id"] if latest_run else ""])
    summary.append(["reconciliation_item_count", len(reconciliation_items)])

    tx_sheet = wb.create_sheet("Transactions")
    _write_sheet(
        tx_sheet,
        [
            "order_id",
            "mandate_ref",
            "receipt_ref",
            "amount",
            "currency",
            "status",
            "vertical",
            "descriptor",
            "audit_index",
            "occurred_at",
            "detail_json",
        ],
        [
            [
                t.get("order_id"),
                t.get("mandate_ref"),
                t.get("receipt_ref"),
                t.get("amount"),
                t.get("currency"),
                t.get("status"),
                t.get("vertical"),
                t.get("descriptor"),
                t.get("audit_index"),
                t.get("occurred_at"),
                t.get("detail_json"),
            ]
            for t in transactions
        ],
    )

    recon_sheet = wb.create_sheet("Reconciliation")
    _write_sheet(
        recon_sheet,
        ["order_id", "mandate_ref", "receipt_ref", "status", "mismatch_reason", "detail_json"],
        [
            [
                item.get("order_id"),
                item.get("mandate_ref"),
                item.get("receipt_ref"),
                item.get("status"),
                item.get("mismatch_reason"),
                item.get("detail_json"),
            ]
            for item in reconciliation_items
        ],
    )

    logs_sheet = wb.create_sheet("OperationLogs")
    _write_sheet(
        logs_sheet,
        ["action", "actor", "merchant_id", "created_at", "detail_json"],
        [
            [
                log.get("action"),
                log.get("actor"),
                log.get("merchant_id"),
                log.get("created_at"),
                log.get("detail_json"),
            ]
            for log in logs
        ],
    )

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def dispute_export_filename(merchant_id: str, generated_at: datetime) -> str:
    return f"dispute_{merchant_id}_{generated_at.strftime('%Y%m%d_%H%M%S')}.xlsx"


async def list_exports(session: AsyncSession, merchant_id: str) -> list[dict[str, Any]]:
    result = await rows(
        session,
        "SELECT * FROM export_jobs WHERE merchant_id=:merchant_id ORDER BY created_at DESC",
        {"merchant_id": merchant_id},
    )
    for row in result:
        row["filters_json"] = loads(row.get("filters_json"))
        row["artifact_summary_json"] = loads(row.get("artifact_summary_json"))
    return result


async def _record_dispute_export_job(
    session: AsyncSession,
    merchant_id: str,
    requested_by: str,
    artifact: dict[str, Any],
    filters: dict[str, Any] | None,
    now: datetime,
) -> dict[str, Any]:
    result = await session.execute(
        text(
            """
            INSERT INTO export_jobs
            (merchant_id, export_type, status, requested_by, filters_json, artifact_summary_json, completed_at)
            VALUES (:merchant_id, 'dispute_bundle', 'COMPLETED', :requested_by, :filters_json, :artifact_summary, :now)
            """
        ),
        {
            "merchant_id": merchant_id,
            "requested_by": requested_by,
            "filters_json": dumps(filters or {}),
            "artifact_summary": dumps(artifact),
            "now": now,
        },
    )
    job_id = result.lastrowid
    await log_operation(
        session,
        "export.dispute_bundle",
        merchant_id=merchant_id,
        actor=requested_by,
        detail={"job_id": job_id, "format": "xlsx"},
    )
    job = await one(session, "SELECT * FROM export_jobs WHERE id=:id", {"id": job_id})
    if job:
        job["filters_json"] = loads(job.get("filters_json"))
        job["artifact_summary_json"] = loads(job.get("artifact_summary_json"))
    return job or {}


async def create_dispute_export(
    session: AsyncSession,
    merchant_id: str,
    requested_by: str = "admin",
    filters: dict[str, Any] | None = None,
) -> dict[str, Any]:
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    bundle = await _gather_dispute_bundle(session, merchant_id)
    artifact = {
        "merchant_id": merchant_id,
        "generated_at": now.isoformat(),
        "transaction_count": len(bundle["transactions"]),
        "log_count": len(bundle["logs"]),
        "reconciliation_run_count": len(bundle["runs"]),
        "reconciliation_item_count": len(bundle["reconciliation_items"]),
        "bundle_format": "xlsx",
        "filename": dispute_export_filename(merchant_id, now),
    }
    return await _record_dispute_export_job(session, merchant_id, requested_by, artifact, filters, now)


async def download_dispute_excel(
    session: AsyncSession,
    merchant_id: str,
    requested_by: str = "admin",
    filters: dict[str, Any] | None = None,
) -> tuple[bytes, str]:
    now = datetime.now(timezone.utc).replace(tzinfo=None)
    bundle = await _gather_dispute_bundle(session, merchant_id)
    content = build_dispute_excel_bytes(merchant_id, bundle, now)
    filename = dispute_export_filename(merchant_id, now)
    artifact = {
        "merchant_id": merchant_id,
        "generated_at": now.isoformat(),
        "transaction_count": len(bundle["transactions"]),
        "log_count": len(bundle["logs"]),
        "reconciliation_run_count": len(bundle["runs"]),
        "reconciliation_item_count": len(bundle["reconciliation_items"]),
        "bundle_format": "xlsx",
        "filename": filename,
    }
    await _record_dispute_export_job(session, merchant_id, requested_by, artifact, filters, now)
    return content, filename
